"""Read a crew manifest out of a CSV, Excel workbook or PDF.

Agents receive manifests in whatever the vessel sends. Previously only CSV was
accepted; Excel and PDF are now handled too.

CSV and Excel are read directly — they are tables, and the column headings just
need matching. PDF is different: the IMO Crew List is a printed form whose
columns are laid out visually, so it goes through Claude the same way
`bill_extraction.py` reads receipts. That keeps PDF support dependency-free
beyond the Anthropic client already in use, and degrades to a clear error when
no API key is configured rather than failing obscurely.

Nothing here touches the database. Callers parse first, show the agent what was
found, and only then persist — so a misread manifest is caught before it lands
on the crew list.
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import os
from datetime import datetime
from typing import List, Optional

from pydantic import BaseModel, Field

logger = logging.getLogger("heyports.crew_manifest")

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
MANIFEST_MODEL = os.getenv("CREW_MANIFEST_MODEL", "claude-haiku-4-5")

CSV_EXTENSIONS = (".csv",)
EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
PDF_EXTENSIONS = (".pdf",)

# Older .xls is a different binary format that openpyxl cannot read.
LEGACY_EXCEL_EXTENSIONS = (".xls",)

DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
    "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d %b %Y",
    "%m/%d/%Y",
)

TRUTHY = {"true", "1", "yes", "y", "checked", "allowed", "eligible"}

COLUMN_ALIASES = {
    "name": ["name", "full name", "crew name", "member name", "family name, given name",
             "family name given name", "surname and name", "crew member"],
    "rank": ["rank", "designation", "role", "position", "rank/rating"],
    "nationality": ["nationality", "country", "nat"],
    "passport_number": ["passport number", "passport", "passport no", "passport_number",
                        "passportno", "passport no.", "passport number "],
    "shore_pass_eligible": ["shore pass allowed or not?", "shore pass allowed", "eligible",
                            "shore_pass_eligible", "allowed", "shore pass allowed or not",
                            "shore pass allowed or not ?", "shore leave allowed"],
    "shore_pass_valid_upto": ["shore pass valid upto", "shore_pass_valid_upto", "valid upto",
                              "validity", "expires", "valid until"],
}


class ParsedCrewRow(BaseModel):
    name: str = Field(description="Family name and given name as printed")
    rank: Optional[str] = Field(default=None, description="Rank or rating, as printed")
    nationality: Optional[str] = Field(default=None, description="Nationality as printed")
    passport_number: Optional[str] = Field(default=None, description="Passport number")
    shore_pass_eligible: bool = False
    shore_pass_valid_upto: Optional[datetime] = None


class ParsedManifest(BaseModel):
    crew: List[ParsedCrewRow] = []
    source: str = ""
    warnings: List[str] = []


class ManifestError(Exception):
    """The file could not be read. The message is shown to the agent."""


# --------------------------------------------------------------------------
# shared helpers
# --------------------------------------------------------------------------

def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _parse_date(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _column_index(headers: List[str], field: str) -> int:
    for alias in COLUMN_ALIASES[field]:
        if alias in headers:
            return headers.index(alias)
    # Fall back to a contains match, which catches "8. Family name, given name"
    # and similar numbered headings on printed forms.
    for alias in COLUMN_ALIASES[field]:
        for idx, header in enumerate(headers):
            if alias and alias in header:
                return idx
    return -1


def _rows_to_manifest(headers: List[str], rows, source: str) -> ParsedManifest:
    headers = [_clean(h).lower() for h in headers]
    idx = {field: _column_index(headers, field) for field in COLUMN_ALIASES}

    if idx["name"] == -1:
        raise ManifestError(
            "Could not find a crew name column. "
            f"Columns found: {', '.join(h for h in headers if h) or 'none'}"
        )

    def cell(row, field):
        i = idx[field]
        return row[i] if i != -1 and len(row) > i else None

    crew, warnings, skipped = [], [], 0
    for row in rows:
        if not row:
            continue
        name = _clean(cell(row, "name"))
        if not name:
            skipped += 1
            continue
        crew.append(ParsedCrewRow(
            name=name,
            rank=_clean(cell(row, "rank")) or None,
            nationality=_clean(cell(row, "nationality")) or None,
            passport_number=(_clean(cell(row, "passport_number")) or None),
            shore_pass_eligible=_clean(cell(row, "shore_pass_eligible")).lower() in TRUTHY,
            shore_pass_valid_upto=_parse_date(cell(row, "shore_pass_valid_upto")),
        ))

    if skipped:
        warnings.append(f"{skipped} row(s) had no crew name and were skipped.")
    for field in ("rank", "passport_number"):
        if idx[field] == -1:
            warnings.append(f"No {field.replace('_', ' ')} column found; left blank.")

    return ParsedManifest(crew=crew, source=source, warnings=warnings)


# --------------------------------------------------------------------------
# per-format readers
# --------------------------------------------------------------------------

def _parse_csv(data: bytes) -> ParsedManifest:
    try:
        decoded = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = data.decode("latin-1")
        except Exception as exc:  # pragma: no cover — defensive
            raise ManifestError("Could not read the CSV. Save it as UTF-8 and try again.") from exc

    reader = csv.reader(io.StringIO(decoded))
    try:
        headers = next(reader)
    except StopIteration:
        raise ManifestError("The CSV file is empty.")
    return _rows_to_manifest(headers, list(reader), "csv")


def _parse_excel(data: bytes) -> ParsedManifest:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — dependency is pinned
        raise ManifestError("Excel support is not installed on the server.") from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ManifestError(
            "Could not read that Excel file. Re-save it as .xlsx and try again."
        ) from exc

    sheet = wb.active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise ManifestError("The Excel sheet is empty.")

    # Printed manifests often carry a title or blank lines above the table, so
    # the header row is not always the first one. Use the first row that names
    # a crew column.
    header_row = 0
    for i, row in enumerate(rows[:15]):
        lowered = [_clean(c).lower() for c in row]
        if _column_index(lowered, "name") != -1:
            header_row = i
            break

    return _rows_to_manifest(rows[header_row], rows[header_row + 1:], "excel")


_PDF_PROMPT = (
    "This is a ship's crew list — often the standard IMO Crew List form. "
    "Extract EVERY crew member in the table, in the order printed. For each: the "
    "full name as printed (family name and given name), the rank or rating exactly "
    "as abbreviated on the form (for example MASTER, CH OFF, 2ND OFF, AB, OS, MTM, "
    "ETO), the nationality, and the passport number. Use the PASSPORT number, not "
    "the seaman's book number, when both are present. Leave a field null if it is "
    "not printed — do not guess or invent crew. Do not include the master's "
    "signature block or any header rows as crew members."
)


def _parse_pdf(data: bytes) -> ParsedManifest:
    if not ANTHROPIC_API_KEY:
        raise ManifestError(
            "PDF manifests need the document reader to be configured on the server "
            "(ANTHROPIC_API_KEY). Upload a CSV or Excel file instead."
        )

    class _PdfCrewList(BaseModel):
        crew: List[ParsedCrewRow] = Field(default_factory=list)

    try:
        import anthropic

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        resp = client.messages.parse(
            model=MANIFEST_MODEL,
            max_tokens=8192,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64.standard_b64encode(data).decode("utf-8"),
                    }},
                    {"type": "text", "text": _PDF_PROMPT},
                ],
            }],
            output_format=_PdfCrewList,
        )
        parsed = resp.parsed_output
    except Exception as exc:
        logger.exception("Crew manifest PDF extraction failed")
        # A rejected key is an operator problem, not a bad scan. Telling the
        # agent to try a clearer photo would send them round in circles.
        detail = f"{type(exc).__name__}: {exc}".lower()
        if "authentication" in detail or "401" in detail or "api key" in detail:
            raise ManifestError(
                "The document reader rejected the server's credentials, so PDF "
                "manifests cannot be read right now. Please report this to "
                "HeyPorts support. Upload a CSV or Excel file in the meantime."
            ) from exc
        if "rate limit" in detail or "429" in detail or "overloaded" in detail:
            raise ManifestError(
                "The document reader is busy. Try again in a moment, or upload a "
                "CSV or Excel file instead."
            ) from exc
        raise ManifestError(
            "Could not read the crew list from that PDF. Try a clearer scan, or "
            "upload a CSV or Excel file instead."
        ) from exc

    if parsed is None or not parsed.crew:
        raise ManifestError(
            "No crew could be read from that PDF. Check it contains a crew list "
            "table, or upload a CSV or Excel file instead."
        )

    crew = [row for row in parsed.crew if _clean(row.name)]
    warnings = []
    if len(crew) != len(parsed.crew):
        warnings.append(f"{len(parsed.crew) - len(crew)} unreadable row(s) were skipped.")
    warnings.append("Read from a PDF — please check the details before saving.")
    return ParsedManifest(crew=crew, source="pdf", warnings=warnings)


# --------------------------------------------------------------------------
# entry point
# --------------------------------------------------------------------------

def parse_manifest(data: bytes, filename: str) -> ParsedManifest:
    """Read a manifest, choosing the reader by file extension."""
    if not data:
        raise ManifestError("The uploaded file is empty.")

    name = (filename or "").lower()
    if name.endswith(CSV_EXTENSIONS):
        return _parse_csv(data)
    if name.endswith(EXCEL_EXTENSIONS):
        return _parse_excel(data)
    if name.endswith(PDF_EXTENSIONS):
        return _parse_pdf(data)
    if name.endswith(LEGACY_EXCEL_EXTENSIONS):
        raise ManifestError(
            "The older .xls format is not supported. Open it and save as .xlsx."
        )
    raise ManifestError("Upload a CSV, Excel (.xlsx) or PDF crew manifest.")
